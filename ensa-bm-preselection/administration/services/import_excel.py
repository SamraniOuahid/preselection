# administration/services/import_excel.py
# Service d'import des notes depuis un fichier Excel

import hashlib
import logging
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db import transaction
from django.utils import timezone

from administration.models import EpreuveEcrite, NoteEcrite
from candidatures.models import Dossier
from users.models import User

logger = logging.getLogger(__name__)


def _colonne_vers_index(col):
    """
    Convertit une lettre de colonne Excel (A, B, C...) en index 0-based,
    ou retourne l'entier directement si c'est déjà un nombre.
    """
    if isinstance(col, int):
        return col
    if isinstance(col, str) and col.isdigit():
        return int(col)
    # Conversion lettre → index (A=0, B=1, ...)
    col = col.upper().strip()
    result = 0
    for char in col:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1  # 0-based


def _nettoyer_cin(valeur):
    """Nettoie un CIN : supprime espaces, met en majuscules."""
    if valeur is None:
        return ''
    return str(valeur).strip().upper()


def _convertir_note(valeur):
    """
    Convertit une valeur en Decimal.
    Gère les virgules (format français) et les points.
    """
    if valeur is None:
        return None
    valeur_str = str(valeur).strip()
    if valeur_str == '' or valeur_str.lower() in ('abs', 'absent', 'a', '-'):
        return None
    # Remplacer virgule par point
    valeur_str = valeur_str.replace(',', '.')
    try:
        return Decimal(valeur_str)
    except (InvalidOperation, ValueError):
        raise ValueError(f"Valeur non numérique : '{valeur}'")


@transaction.atomic
def importer_notes_excel(fichier, epreuve_id, colonne_cin='A',
                         colonne_note='B', ligne_debut=2):
    """
    Importe les notes d'un fichier Excel pour une épreuve écrite.

    Args:
        fichier: Fichier Excel uploadé (InMemoryUploadedFile)
        epreuve_id: UUID de l'EpreuveEcrite
        colonne_cin: Lettre ou index de la colonne CIN (ex: 'A' ou 0)
        colonne_note: Lettre ou index de la colonne note (ex: 'B' ou 1)
        ligne_debut: Première ligne de données (défaut 2, ligne 1 = en-têtes)

    Returns:
        dict avec le rapport complet d'import
    """
    rapport = {
        'succes': False,
        'total_lignes': 0,
        'importees': 0,
        'erreurs': [],
        'non_trouves': [],
        'nb_admis': 0,
        'nb_recales': 0,
        'nb_absents': 0,
    }

    # Récupérer l'épreuve
    try:
        epreuve = EpreuveEcrite.objects.select_related('filiere').get(id=epreuve_id)
    except EpreuveEcrite.DoesNotExist:
        rapport['erreurs'].append({
            'ligne': 0, 'cin': '', 'valeur': '',
            'message': f"Épreuve {epreuve_id} introuvable."
        })
        return rapport

    # ÉTAPE 1 — Lire le fichier Excel
    try:
        # Calculer le hash MD5 pour traçabilité
        fichier.seek(0)
        file_content = fichier.read()
        file_hash = hashlib.md5(file_content).hexdigest()
        fichier.seek(0)

        wb = openpyxl.load_workbook(fichier, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        rapport['erreurs'].append({
            'ligne': 0, 'cin': '', 'valeur': '',
            'message': f"Impossible de lire le fichier Excel : {e}"
        })
        return rapport

    # Convertir les indices de colonnes
    idx_cin = _colonne_vers_index(colonne_cin)
    idx_note = _colonne_vers_index(colonne_note)

    # ÉTAPE 2 — Lire les lignes
    lignes = list(ws.iter_rows(min_row=ligne_debut, values_only=True))
    if not lignes:
        rapport['erreurs'].append({
            'ligne': 0, 'cin': '', 'valeur': '',
            'message': "Le fichier Excel est vide (aucune ligne de données)."
        })
        wb.close()
        return rapport

    rapport['total_lignes'] = len(lignes)

    # ÉTAPE 3 — Traiter chaque ligne
    for i, row in enumerate(lignes):
        num_ligne = ligne_debut + i

        # Vérifier que la ligne a assez de colonnes
        if row is None or len(row) <= max(idx_cin, idx_note):
            rapport['erreurs'].append({
                'ligne': num_ligne, 'cin': '', 'valeur': '',
                'message': "Ligne incomplète (pas assez de colonnes)."
            })
            continue

        # Lire le CIN
        cin_brut = row[idx_cin]
        cin = _nettoyer_cin(cin_brut)
        if not cin:
            rapport['erreurs'].append({
                'ligne': num_ligne, 'cin': '', 'valeur': str(cin_brut),
                'message': "CIN vide ou invalide."
            })
            continue

        # Lire et convertir la note
        note_brute = row[idx_note]
        try:
            note = _convertir_note(note_brute)
        except ValueError as e:
            rapport['erreurs'].append({
                'ligne': num_ligne, 'cin': cin,
                'valeur': str(note_brute),
                'message': str(e)
            })
            continue

        # Valider la plage de la note
        if note is not None:
            if note < 0 or note > epreuve.note_sur:
                rapport['erreurs'].append({
                    'ligne': num_ligne, 'cin': cin,
                    'valeur': str(note_brute),
                    'message': f"Note hors plage (0 — {epreuve.note_sur})."
                })
                continue

        # Chercher le User avec ce CIN
        try:
            user = User.objects.get(cin=cin)
        except User.DoesNotExist:
            rapport['non_trouves'].append({
                'ligne': num_ligne, 'cin': cin
            })
            continue

        # Chercher le Dossier PRESELECTIONNE pour cette filière
        try:
            dossier = Dossier.objects.get(
                candidat__user=user,
                filiere=epreuve.filiere,
                statut__in=[
                    Dossier.Statut.PRESELECTIONNE,
                    Dossier.Statut.ADMIS_FINAL,
                    Dossier.Statut.RECALE_FINAL,
                    Dossier.Statut.ABSENT_ECRIT,
                ]
            )
        except Dossier.DoesNotExist:
            rapport['non_trouves'].append({
                'ligne': num_ligne, 'cin': cin
            })
            continue

        # Créer ou mettre à jour la NoteEcrite
        note_ecrite, _ = NoteEcrite.objects.update_or_create(
            epreuve=epreuve,
            dossier=dossier,
            defaults={
                'note': note,
                'ligne_excel': num_ligne,
                'import_hash': file_hash,
                'note_importee_le': timezone.now(),
            }
        )
        note_ecrite.calculer_resultat()
        rapport['importees'] += 1

    wb.close()

    # ÉTAPE 4 — Mise à jour du classement
    recalculer_classement_final(epreuve)

    # ÉTAPE 5 — Mettre à jour le statut
    epreuve.statut = EpreuveEcrite.Statut.NOTES_IMPORTEES
    epreuve.save()

    # Compter les résultats
    rapport['nb_admis'] = epreuve.notes.filter(
        resultat=NoteEcrite.Resultat.ADMIS
    ).count()
    rapport['nb_recales'] = epreuve.notes.filter(
        resultat=NoteEcrite.Resultat.RECALE
    ).count()
    rapport['nb_absents'] = epreuve.notes.filter(
        resultat=NoteEcrite.Resultat.ABSENT
    ).count()
    rapport['succes'] = True

    logger.info(
        f"Import terminé pour {epreuve.nom}: "
        f"{rapport['importees']} notes importées, "
        f"{len(rapport['erreurs'])} erreurs, "
        f"{len(rapport['non_trouves'])} non trouvés"
    )

    return rapport


def recalculer_classement_final(epreuve):
    """
    Recalcule le classement final et met à jour les statuts des dossiers.

    Pour tous les admis :
      - Trier par note DESC, puis par score dossier DESC
      - Assigner rang_final 1, 2, 3...
      - Calculer le score_final combiné
      - Mettre à jour le statut du dossier
    """
    # Récupérer toutes les notes avec dossier lié
    notes = epreuve.notes.select_related(
        'dossier'
    ).all()

    # Mettre à jour les statuts des dossiers
    for note in notes:
        dossier = note.dossier
        if note.resultat == NoteEcrite.Resultat.ADMIS:
            # Calculer score_final
            score_dossier = float(dossier.score or 0)
            note_ecrite = float(note.note or 0)
            # Normaliser la note sur 20
            note_normalisee = (note_ecrite / float(epreuve.note_sur)) * 20
            # Pondération : 40% dossier + 60% épreuve (configurable via coefficient)
            coeff_ecrit = float(epreuve.coefficient)
            score_final = (score_dossier * 0.4) + (note_normalisee * 0.6 * coeff_ecrit)
            dossier.score_final = Decimal(str(round(score_final, 2)))
            dossier.statut = Dossier.Statut.ADMIS_FINAL
            dossier.save()
        elif note.resultat == NoteEcrite.Resultat.RECALE:
            dossier.statut = Dossier.Statut.RECALE_FINAL
            dossier.save()
        elif note.resultat == NoteEcrite.Resultat.ABSENT:
            dossier.statut = Dossier.Statut.ABSENT_ECRIT
            dossier.save()

    # Calculer le classement des admis
    admis_notes = epreuve.notes.filter(
        resultat=NoteEcrite.Resultat.ADMIS
    ).select_related('dossier').order_by('-note', '-dossier__score')

    for rang, note in enumerate(admis_notes, start=1):
        note.rang_final = rang
        note.save(update_fields=['rang_final'])
        note.dossier.rang_final = rang
        note.dossier.save(update_fields=['rang_final'])


def recalculer_apres_changement_seuil(epreuve):
    """
    Recalcule tous les résultats après un changement de seuil d'admission.

    Returns:
        dict avec anciens_admis, nouveaux_admis, difference
    """
    # Compter les admis actuels avant recalcul
    anciens_admis = epreuve.notes.filter(
        resultat=NoteEcrite.Resultat.ADMIS
    ).count()

    # Recalculer le résultat de chaque note
    for note in epreuve.notes.all():
        note.calculer_resultat()

    # Recalculer le classement
    recalculer_classement_final(epreuve)

    # Compter les nouveaux admis
    nouveaux_admis = epreuve.notes.filter(
        resultat=NoteEcrite.Resultat.ADMIS
    ).count()

    return {
        'anciens_admis': anciens_admis,
        'nouveaux_admis': nouveaux_admis,
        'difference': nouveaux_admis - anciens_admis,
    }


def previsualiser_excel(fichier, colonne_cin='A', colonne_note='B', ligne_debut=2):
    """
    Retourne un aperçu des premières lignes du fichier Excel.

    Returns:
        dict avec colonnes détectées et les 5 premières lignes
    """
    try:
        wb = openpyxl.load_workbook(fichier, read_only=True, data_only=True)
        ws = wb.active

        # En-têtes (ligne 1)
        headers = []
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if header_row:
            headers = [str(h) if h else f"Colonne {chr(65+i)}"
                       for i, h in enumerate(header_row[0])]

        # Premières lignes de données
        apercu = []
        for row in ws.iter_rows(min_row=ligne_debut,
                                max_row=ligne_debut + 4,
                                values_only=True):
            apercu.append([str(v) if v is not None else '' for v in row])

        # Nombre total de lignes
        total_lignes = sum(1 for _ in ws.iter_rows(min_row=ligne_debut, values_only=True))

        wb.close()

        return {
            'succes': True,
            'headers': headers,
            'apercu': apercu,
            'total_lignes': total_lignes,
            'nb_colonnes': len(headers),
        }
    except Exception as e:
        return {
            'succes': False,
            'message': f"Erreur de lecture : {e}",
        }


def generer_template_excel():
    """
    Génère un fichier Excel template pour l'import des notes.

    Returns:
        openpyxl.Workbook
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notes"

    # Styles
    header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # En-têtes
    headers = ['CIN', 'Note']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Exemples fictifs
    exemples = [
        ('AB123456', 14.5),
        ('CD789012', 16.0),
        ('EF345678', 11.25),
        ('GH901234', 8.0),
        ('IJ567890', None),  # Absent
    ]
    for row_idx, (cin, note) in enumerate(exemples, 2):
        ws.cell(row=row_idx, column=1, value=cin).border = border
        note_cell = ws.cell(row=row_idx, column=2, value=note if note else 'ABS')
        note_cell.border = border

    # Ajuster largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    # Feuille Instructions
    ws2 = wb.create_sheet(title="Instructions")
    instructions = [
        "INSTRUCTIONS POUR L'IMPORT DES NOTES",
        "",
        "Format du fichier :",
        "  - Colonne A : CIN du candidat (ex: AB123456)",
        "  - Colonne B : Note obtenue (nombre décimal)",
        "",
        "Format CIN accepté :",
        "  - Lettres et chiffres, sans espaces",
        "  - Exemple : AB123456, CD789012",
        "  - Le CIN doit correspondre exactement à celui enregistré",
        "",
        "Format Note :",
        "  - Nombre décimal avec point ou virgule (ex: 14.5 ou 14,5)",
        "  - Pour un absent : laisser vide ou écrire 'ABS'",
        "  - La note doit être comprise entre 0 et le barème de l'épreuve",
        "",
        "Remarques :",
        "  - La première ligne (en-têtes) est ignorée",
        "  - Les CIN non trouvés dans la base seront signalés",
        "  - Les notes hors plage seront signalées comme erreurs",
    ]
    for row_idx, text in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
    ws2.column_dimensions['A'].width = 70

    return wb
