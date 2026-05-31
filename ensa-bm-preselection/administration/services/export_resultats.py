# administration/services/export_resultats.py
# Service d'export des résultats d'une épreuve écrite au format Excel

import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Min, Max, Avg, Count

from administration.models import EpreuveEcrite, NoteEcrite

logger = logging.getLogger(__name__)


def _style_header(ws, row, fill_color, nb_cols):
    """Applique le style aux en-têtes."""
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    for col in range(1, nb_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def _style_row(ws, row, nb_cols, is_even):
    """Applique l'alternance de couleurs sur les lignes de données."""
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid") if is_even else None
    for col in range(1, nb_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fill:
            cell.fill = fill


def _auto_width(ws, nb_cols, start_row=1, end_row=None):
    """Auto-ajuste la largeur des colonnes."""
    if end_row is None:
        end_row = ws.max_row
    for col in range(1, nb_cols + 1):
        max_length = 0
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max_length + 4, 40)
        ws.column_dimensions[get_column_letter(col)].width = adjusted


def exporter_resultats_excel(epreuve_id):
    """
    Génère un fichier Excel formaté avec les résultats d'une épreuve.

    Returns:
        HttpResponse avec le fichier Excel en pièce jointe
    """
    epreuve = EpreuveEcrite.objects.select_related('filiere').get(id=epreuve_id)

    wb = openpyxl.Workbook()

    # ── Colonnes communes ──
    colonnes = [
        'Rang', 'CIN', 'Nom', 'Prénom', 'Filière',
        'Note Dossier', 'Note Écrite', 'Score Final', 'Résultat'
    ]
    nb_cols = len(colonnes)

    # ── Feuille 1 — Admis ──
    ws_admis = wb.active
    ws_admis.title = "Admis"
    ws_admis.append(colonnes)
    _style_header(ws_admis, 1, "27AE60", nb_cols)

    admis = epreuve.notes.filter(
        resultat=NoteEcrite.Resultat.ADMIS
    ).select_related(
        'dossier', 'dossier__candidat', 'dossier__candidat__user',
        'dossier__filiere'
    ).order_by('rang_final')

    for i, note in enumerate(admis):
        d = note.dossier
        c = d.candidat
        row_data = [
            note.rang_final or i + 1,
            c.user.cin,
            c.nom,
            c.prenom,
            d.filiere.nom,
            float(d.score or 0),
            float(note.note or 0),
            float(d.score_final or 0),
            'Admis'
        ]
        ws_admis.append(row_data)
        _style_row(ws_admis, i + 2, nb_cols, i % 2 == 0)

    _auto_width(ws_admis, nb_cols)

    # ── Feuille 2 — Recalés ──
    ws_recales = wb.create_sheet(title="Recalés")
    colonnes_recales = [
        '#', 'CIN', 'Nom', 'Prénom', 'Filière',
        'Note Dossier', 'Note Écrite', 'Score Final', 'Résultat'
    ]
    ws_recales.append(colonnes_recales)
    _style_header(ws_recales, 1, "C0392B", nb_cols)

    recales = epreuve.notes.filter(
        resultat=NoteEcrite.Resultat.RECALE
    ).select_related(
        'dossier', 'dossier__candidat', 'dossier__candidat__user',
        'dossier__filiere'
    ).order_by('-note')

    for i, note in enumerate(recales):
        d = note.dossier
        c = d.candidat
        row_data = [
            i + 1,
            c.user.cin,
            c.nom,
            c.prenom,
            d.filiere.nom,
            float(d.score or 0),
            float(note.note or 0),
            float(d.score_final or 0),
            'Recalé'
        ]
        ws_recales.append(row_data)
        _style_row(ws_recales, i + 2, nb_cols, i % 2 == 0)

    _auto_width(ws_recales, nb_cols)

    # ── Feuille 3 — Statistiques ──
    ws_stats = wb.create_sheet(title="Statistiques")

    stats = epreuve.notes.filter(note__isnull=False).aggregate(
        note_min=Min('note'),
        note_max=Max('note'),
        note_moyenne=Avg('note'),
    )

    total_candidats = epreuve.notes.count()
    nb_admis = epreuve.notes.filter(resultat=NoteEcrite.Resultat.ADMIS).count()
    nb_recales = epreuve.notes.filter(resultat=NoteEcrite.Resultat.RECALE).count()
    nb_absents = epreuve.notes.filter(resultat=NoteEcrite.Resultat.ABSENT).count()
    taux_admission = round((nb_admis / total_candidats * 100), 1) if total_candidats > 0 else 0

    # Style de la feuille statistiques
    title_font = Font(bold=True, size=14, color="1B3A6B")
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)

    stat_data = [
        ("STATISTIQUES DE L'ÉPREUVE", ''),
        ('', ''),
        ("Épreuve", epreuve.nom),
        ("Filière", epreuve.filiere.nom),
        ("Date de l'épreuve", str(epreuve.date_epreuve or 'Non définie')),
        ('', ''),
        ("Total candidats", total_candidats),
        ("Nombre d'admis", nb_admis),
        ("Taux d'admission", f"{taux_admission}%"),
        ("Nombre de recalés", nb_recales),
        ("Nombre d'absents", nb_absents),
        ('', ''),
        ("Note minimale", float(stats['note_min'] or 0)),
        ("Note maximale", float(stats['note_max'] or 0)),
        ("Note moyenne", round(float(stats['note_moyenne'] or 0), 2)),
        ("Seuil d'admission", float(epreuve.seuil_admission)),
        ("Barème", f"/{float(epreuve.note_sur)}"),
        ('', ''),
        ("Date de génération", datetime.now().strftime('%d/%m/%Y à %H:%M')),
    ]

    for row_idx, (label, value) in enumerate(stat_data, 1):
        cell_label = ws_stats.cell(row=row_idx, column=1, value=label)
        cell_value = ws_stats.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_label.font = title_font
        else:
            cell_label.font = label_font
            cell_value.font = value_font

    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 40

    # Générer la réponse HTTP
    date_str = datetime.now().strftime('%Y%m%d')
    filename = f"Resultats_{epreuve.filiere.code}_{date_str}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    logger.info(f"Export résultats généré : {filename}")
    return response
