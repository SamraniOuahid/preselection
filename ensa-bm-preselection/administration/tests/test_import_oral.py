# administration/tests/test_import_oral.py
"""
Tests de robustesse pour l'import Excel/CSV des admis à l'oral.
Couverture :
  a) Import valide xlsx → candidats convoqués + horaires depuis filiere.date_oral
  b) Import valide csv → candidat convoqué
  c) Fichier sans contenu → 400
  d) Colonnes manquantes → 400
  e) Format non supporté → 400
  f) Candidat hors filière → ignoré (réponse 200, ignores=1)
  g) Statut incompatible → ignoré
  h) Doublon dans le fichier → 2e occurrence ignorée
"""

import datetime
import io
import csv as csv_module

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from administration.models import Filiere, EpreuveOrale, ConvocationOrale
from candidatures.models import Dossier
from users.models import User, Candidat


# ─────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────

def _make_xlsx(colonnes: list, lignes: list[list]) -> SimpleUploadedFile:
    """Génère un SimpleUploadedFile xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(colonnes)
    for row in lignes:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return SimpleUploadedFile('import.xlsx', buf.read(),
                              content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _make_csv(colonnes: list, lignes: list[list]) -> SimpleUploadedFile:
    """Génère un SimpleUploadedFile csv."""
    buf = io.StringIO()
    writer = csv_module.writer(buf)
    writer.writerow(colonnes)
    for row in lignes:
        writer.writerow(row)
    return SimpleUploadedFile('import.csv', buf.getvalue().encode('utf-8'), content_type='text/csv')


# ─────────────────────────────────────────────────────────────────
# Fixture de base
# ─────────────────────────────────────────────────────────────────

class BaseImportOralTest(TestCase):

    def setUp(self):
        self.client = APIClient()

        self.admin = User.objects.create_superuser(
            email='admin@ensa.ma', cin='ADMIN01', password='pass'
        )
        self.client.force_authenticate(user=self.admin)

        self.filiere = Filiere.objects.create(
            nom='Test GI', code='GI-TST', niveau=Filiere.Niveau.BAC2,
            date_oral=timezone.make_aware(datetime.datetime(2025, 9, 15, 9, 0)),
            lieu_oral='Amphi A — ENSA BM',
        )
        self.autre_filiere = Filiere.objects.create(
            nom='Autre', code='AUT', niveau=Filiere.Niveau.BAC2
        )

        self.epreuve = EpreuveOrale.objects.create(
            filiere=self.filiere,
            nom='Oral GI — 2025',
            duree_minutes=20,
            statut=EpreuveOrale.Statut.PLANIFIEE,
        )

        self.dossier1 = self._dossier('C001', 'A1001', 'Alami',   'Karim',   self.filiere)
        self.dossier2 = self._dossier('C002', 'A1002', 'Benali',  'Sara',    self.filiere)
        self.dossier3 = self._dossier('C003', 'A1003', 'Chraibi', 'Youssef', self.filiere)
        self.dossier_autre    = self._dossier('C099', 'A9999', 'Dupont', 'Jean', self.autre_filiere)
        self.dossier_brouillon = self._dossier('C004', 'A1004', 'El Fassi', 'Omar', self.filiere,
                                               statut=Dossier.Statut.BROUILLON)

    def _dossier(self, cin, cne, nom, prenom, filiere, statut=Dossier.Statut.ADMIS_FINAL):
        u = User.objects.create_user(email=f'{cin}@test.ma', cin=cin, password='pass')
        c = Candidat.objects.create(user=u, nom=nom, prenom=prenom)
        return Dossier.objects.create(
            candidat=c, filiere=filiere, statut=statut,
            diplome_obtenu='DUT Info', etablissement_origine='EST',
            cne=cne, code_massar=cne,
        )

    def _url(self):
        return f'/api/epreuves-oral/{self.epreuve.id}/importer-admis-oral/'


# ─────────────────────────────────────────────────────────────────
# a) Import valide xlsx
# ─────────────────────────────────────────────────────────────────

class TestImportValideXlsx(BaseImportOralTest):

    def test_deux_candidats_convoques(self):
        f = _make_xlsx(['cne'], [['A1001'], ['A1002']])
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')

        self.assertEqual(res.status_code, status.HTTP_200_OK, res.data)
        self.assertEqual(res.data['traites'], 2)
        self.assertEqual(res.data['convoques'], 2)
        self.assertEqual(res.data['ignores'], 0)

        self.dossier1.refresh_from_db()
        self.dossier2.refresh_from_db()
        self.assertEqual(self.dossier1.statut, Dossier.Statut.CONVOQUE_ORAL)
        self.assertEqual(self.dossier2.statut, Dossier.Statut.CONVOQUE_ORAL)

        self.epreuve.refresh_from_db()
        self.assertEqual(self.epreuve.statut, EpreuveOrale.Statut.EN_COURS)

    def test_horaires_calcules_depuis_filiere_date_oral(self):
        f = _make_xlsx(['cne'], [['A1001'], ['A1002'], ['A1003']])
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK, res.data)

        convs = list(ConvocationOrale.objects.filter(
            epreuve_oral=self.epreuve
        ).order_by('numero_passage'))
        self.assertEqual(len(convs), 3)
        # heure_debut = 09:00 (filiere.date_oral), duree = 20 min
        self.assertEqual(convs[0].heure_passage, datetime.time(9, 0))
        self.assertEqual(convs[1].heure_passage, datetime.time(9, 20))
        self.assertEqual(convs[2].heure_passage, datetime.time(9, 40))


# ─────────────────────────────────────────────────────────────────
# b) Import valide csv
# ─────────────────────────────────────────────────────────────────

class TestImportValideCSV(BaseImportOralTest):

    def test_un_candidat_convoque(self):
        f = _make_csv(['cne'], [['A1003']])
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')

        self.assertEqual(res.status_code, status.HTTP_200_OK, res.data)
        self.assertEqual(res.data['convoques'], 1)

        self.dossier3.refresh_from_db()
        self.assertEqual(self.dossier3.statut, Dossier.Statut.CONVOQUE_ORAL)

    def test_csv_colonne_code_massar(self):
        f = _make_csv(['code_massar'], [['A1001']])
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK, res.data)
        self.assertEqual(res.data['convoques'], 1)


# ─────────────────────────────────────────────────────────────────
# c) Fichier sans contenu → 400
# ─────────────────────────────────────────────────────────────────

class TestImportFichierVide(BaseImportOralTest):

    def test_xlsx_vide_retourne_400(self):
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        f = SimpleUploadedFile('vide.xlsx', buf.read(),
                               content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_csv_entete_seule_retourne_400(self):
        f = SimpleUploadedFile('empty.csv', b'cne\n', content_type='text/csv')
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_sans_fichier_retourne_400(self):
        res = self.client.post(self._url(), {}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('fichier', res.data['error'].lower())


# ─────────────────────────────────────────────────────────────────
# d) Colonnes manquantes → 400
# ─────────────────────────────────────────────────────────────────

class TestImportColonnesManquantes(BaseImportOralTest):

    def test_colonne_invalide_retourne_400(self):
        f = _make_xlsx(['nom', 'prenom'], [['Alami', 'Karim']])
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        # Le message doit mentionner les colonnes attendues
        err = str(res.data.get('error', '')).lower()
        self.assertTrue(
            'cne' in err or 'code_massar' in err or 'colonne' in err,
            f"Message inattendu : {err}"
        )


# ─────────────────────────────────────────────────────────────────
# e) Format non supporté → 400
# ─────────────────────────────────────────────────────────────────

class TestImportFormatInvalide(BaseImportOralTest):

    def test_pdf_retourne_400(self):
        f = SimpleUploadedFile('doc.pdf', b'%PDF', content_type='application/pdf')
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('format', res.data['error'].lower())


# ─────────────────────────────────────────────────────────────────
# f) Candidat hors filière → ignoré (200, ignores=1)
# ─────────────────────────────────────────────────────────────────

class TestImportCandidatHorsFiliere(BaseImportOralTest):

    def test_cne_autre_filiere_ignore(self):
        # A9999 appartient à autre_filiere — doit être ignoré proprement
        f = _make_xlsx(['cne'], [['A1001'], ['A9999']])
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')

        self.assertEqual(res.status_code, status.HTTP_200_OK, res.data)
        self.assertEqual(res.data['traites'], 2)
        self.assertEqual(res.data['convoques'], 1)
        self.assertEqual(res.data['ignores'], 1)
        self.assertEqual(len(res.data['erreurs']), 1)

        self.dossier_autre.refresh_from_db()
        self.assertNotEqual(self.dossier_autre.statut, Dossier.Statut.CONVOQUE_ORAL)


# ─────────────────────────────────────────────────────────────────
# g) Statut incompatible → ignoré
# ─────────────────────────────────────────────────────────────────

class TestImportStatutIncompatible(BaseImportOralTest):

    def test_brouillon_ignore(self):
        # A1004 est BROUILLON, ne figure pas dans statuts_importables
        f = _make_xlsx(['cne'], [['A1001'], ['A1004']])
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')

        self.assertEqual(res.status_code, status.HTTP_200_OK, res.data)
        self.assertEqual(res.data['convoques'], 1)
        self.assertEqual(res.data['ignores'], 1)

        self.dossier_brouillon.refresh_from_db()
        self.assertNotEqual(self.dossier_brouillon.statut, Dossier.Statut.CONVOQUE_ORAL)


# ─────────────────────────────────────────────────────────────────
# h) Doublon dans le même fichier → 2e occurrence ignorée
# ─────────────────────────────────────────────────────────────────

class TestImportDoublons(BaseImportOralTest):

    def test_doublon_ignore(self):
        f = _make_xlsx(['cne'], [['A1001'], ['A1001']])
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')

        self.assertEqual(res.status_code, status.HTTP_200_OK, res.data)
        self.assertEqual(res.data['traites'], 2)
        self.assertEqual(res.data['convoques'], 1)
        self.assertEqual(res.data['ignores'], 1)

        count = ConvocationOrale.objects.filter(
            epreuve_oral=self.epreuve, dossier=self.dossier1
        ).count()
        self.assertEqual(count, 1)


# ─────────────────────────────────────────────────────────────────
# i) Fichier avec en-tête précédé de lignes de titre/bannière
# ─────────────────────────────────────────────────────────────────

def _make_xlsx_with_banner(banner: list, colonnes: list, lignes: list[list]) -> SimpleUploadedFile:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(banner)
    ws.append([])  # Ligne vide
    ws.append(colonnes)
    for row in lignes:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return SimpleUploadedFile('import_banner.xlsx', buf.read(),
                              content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _make_csv_with_banner(banner: list, colonnes: list, lignes: list[list]) -> SimpleUploadedFile:
    buf = io.StringIO()
    writer = csv_module.writer(buf)
    writer.writerow(banner)
    writer.writerow([])  # Ligne vide
    writer.writerow(colonnes)
    for row in lignes:
        writer.writerow(row)
    return SimpleUploadedFile('import_banner.csv', buf.getvalue().encode('utf-8'), content_type='text/csv')


class TestImportWithBanner(BaseImportOralTest):

    def test_import_xlsx_avec_banniere(self):
        f = _make_xlsx_with_banner(
            ['ensa béni mellal — admis à l épreuve orale — filière tdi', '', '', ''],
            ['cne'],
            [['A1001'], ['A1002']]
        )
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK, res.data)
        self.assertEqual(res.data['traites'], 2)
        self.assertEqual(res.data['convoques'], 2)
        self.assertEqual(res.data['ignores'], 0)

        self.dossier1.refresh_from_db()
        self.dossier2.refresh_from_db()
        self.assertEqual(self.dossier1.statut, Dossier.Statut.CONVOQUE_ORAL)
        self.assertEqual(self.dossier2.statut, Dossier.Statut.CONVOQUE_ORAL)

    def test_import_csv_avec_banniere(self):
        f = _make_csv_with_banner(
            ['ensa béni mellal — admis à l épreuve orale — filière tdi', '', '', ''],
            ['code_massar'],
            [['A1001']]
        )
        res = self.client.post(self._url(), {'fichier': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK, res.data)
        self.assertEqual(res.data['traites'], 1)
        self.assertEqual(res.data['convoques'], 1)
        self.assertEqual(res.data['ignores'], 0)

        self.dossier1.refresh_from_db()
        self.assertEqual(self.dossier1.statut, Dossier.Statut.CONVOQUE_ORAL)

